import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

JSON_PATH = r"C:\Claude\gita-concept-kg\_extracted\clean_verses_700.json"
XLSX_PATH = r"C:\Claude\gita-concept-kg\Bhagavad_Gita_All_Verses_CLEAN.xlsx"

def build_xlsx():
    print(f"Loading {JSON_PATH}...")
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    print(f"Loaded {len(data)} verses.")

    # Create workbook
    wb = openpyxl.Workbook()
    
    # ---------------- 1. README Sheet ----------------
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.views.sheetView[0].showGridLines = True
    
    # Header styles for README
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    header_font = Font(name="Segoe UI", size=11, bold=True)
    body_font = Font(name="Segoe UI", size=11)
    bold_body_font = Font(name="Segoe UI", size=11, bold=True)
    
    # Write Readme content
    ws_readme.cell(row=2, column=2, value="Bhagavad Gītā — Complete Kannada Commentary").font = title_font
    ws_readme.cell(row=3, column=2, value="Sanskrit Ślokas + Kannada Commentary by Shri Bannanje Govindacharya").font = Font(name="Segoe UI", size=12, italic=True, color="595959")
    
    ws_readme.cell(row=5, column=2, value="This workbook contains the complete set of verses from the Bhagavad Gita as compiled in Shri Bannanje Govindacharya's edition.").font = body_font
    ws_readme.cell(row=6, column=2, value="Because Shri Bannanje's division and numbering of certain verses differ from conventional Bhagavad Gita editions, this dataset maps both, keeping the conventional numbering structure as the key reference.").font = body_font
    
    # Legend Table
    ws_readme.cell(row=9, column=2, value="COLOR LEGEND & VERSE STATUS:").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
    
    legend_headers = ["Status Flag", "Background Color", "Description"]
    for col_idx, h in enumerate(legend_headers, start=2):
        cell = ws_readme.cell(row=10, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = Border(bottom=Side(style="medium"), top=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
        
    legend_data = [
        ("clean", "White (No Fill)", "Fully clean OCR Unicode text verified against the printed edition."),
        ("screenshot_patch", "Soft Green", "Manually transcribed and patched from printed book page screenshots shared by the user."),
        ("auto_extracted", "Soft Blue", "Heuristically extracted and split between Shloka and Meaning columns (some columns splits may have minor alignment overlaps)."),
        ("phantom_disregard", "Soft Gray", "Phantom rows — verses that do not exist under Bannanje Govindacharya's numbering. Retained as placeholders for conventional indexing.")
    ]
    
    fills = {
        "clean": PatternFill(fill_type=None),
        "screenshot_patch": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "auto_extracted": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        "phantom_disregard": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    }
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row_idx, (status, color_name, desc) in enumerate(legend_data, start=11):
        c1 = ws_readme.cell(row=row_idx, column=2, value=status)
        c2 = ws_readme.cell(row=row_idx, column=3, value=color_name)
        c3 = ws_readme.cell(row=row_idx, column=4, value=desc)
        
        c1.font = bold_body_font
        c2.font = body_font
        c3.font = body_font
        
        # Apply border
        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border
        
        # Apply fill to the background column
        c2.fill = fills[status]
        
    ws_readme.column_dimensions["B"].width = 25
    ws_readme.column_dimensions["C"].width = 22
    ws_readme.column_dimensions["D"].width = 85
    
    # ---------------- Helper function for Data Sheets ----------------
    def write_sheet_data(ws, items):
        ws.views.sheetView[0].showGridLines = True
        ws.freeze_panes = "A2"
        
        headers = ["Chapter", "Verse", "Ref", "Sanskrit Shloka", "Kannada Meaning & Commentary", "Source Page", "Status"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(name="Segoe UI", size=11, bold=True, color="000000")
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.border = Border(bottom=Side(style="medium"), left=Side(style="thin"), right=Side(style="thin"))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for r_idx, item in enumerate(items, start=2):
            c_val = item.get("chapter", "")
            v_val = item.get("verse", "")
            ref_val = item.get("ref", "")
            shloka_val = item.get("shloka", "")
            meaning_val = item.get("meaning", "")
            src_val = item.get("source_page", "")
            status_val = item.get("status", "")
            
            row_cells = [
                ws.cell(row=r_idx, column=1, value=c_val),
                ws.cell(row=r_idx, column=2, value=v_val),
                ws.cell(row=r_idx, column=3, value=ref_val),
                ws.cell(row=r_idx, column=4, value=shloka_val),
                ws.cell(row=r_idx, column=5, value=meaning_val),
                ws.cell(row=r_idx, column=6, value=src_val),
                ws.cell(row=r_idx, column=7, value=status_val),
            ]
            
            fill = fills.get(status_val, PatternFill(fill_type=None))
            
            for idx, cell in enumerate(row_cells, start=1):
                cell.font = Font(name="Segoe UI", size=10)
                cell.border = thin_border
                cell.alignment = Alignment(
                    vertical="top",
                    horizontal="center" if idx in [1, 2, 3, 6, 7] else "left",
                    wrap_text=True if idx in [4, 5] else False
                )
                if fill.fill_type:
                    cell.fill = fill
                    
        # Column dimensions
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 42
        ws.column_dimensions["E"].width = 85
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 18
        
        ws.row_dimensions[1].height = 25
        
    # ---------------- 2. All Verses Sheet ----------------
    print("Writing 'All Verses' sheet...")
    ws_all = wb.create_sheet("All Verses")
    write_sheet_data(ws_all, data)
    
    # ---------------- 3. Chapter 1 to 18 Sheets ----------------
    for ch in range(1, 19):
        print(f"Writing 'Chapter {ch}' sheet...")
        ws_ch = wb.create_sheet(f"Chapter {ch}")
        ch_items = [x for x in data if x.get("chapter") == ch]
        write_sheet_data(ws_ch, ch_items)
        
    # Save Workbook
    print(f"Saving workbook to {XLSX_PATH}...")
    wb.save(XLSX_PATH)
    print("Excel file successfully rebuilt!")

if __name__ == "__main__":
    build_xlsx()
